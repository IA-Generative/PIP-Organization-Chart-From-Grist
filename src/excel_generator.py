from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import os
import re
import threading
from dataclasses import dataclass
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .model_builder import BuiltModel, EpicModel

DEFAULT_SCW_BASE_URL = "https://api.scaleway.ai/a9158aac-8404-46ea-8bf5-1ca048cd6ab4/v1"
DEFAULT_EXCEL_MODEL = "ossgpt"
ROBOT_TAG = "(🤖)"
SYNTH_MAX_WORDS = 96
INTENTION_MAX_WORDS = 56
JUSTIF_MAX_WORDS = 72
FEATURE_PROPOSAL_MAX_WORDS = 40


@dataclass
class EpicEntry:
    team_name: str
    epic: EpicModel


@dataclass
class EpicAnalysis:
    synthese_courte: str
    intention_courte: str
    couverture_features: str
    justification: str
    proposals: List[Tuple[str, str]]
    ai_used: bool


def _llm_enabled() -> bool:
    return os.getenv("ENABLE_LLM", "0") == "1" and bool(os.getenv("SCW_SECRET_KEY_LLM"))


def _excel_llm_max_workers() -> int:
    raw = os.getenv("EXCEL_LLM_MAX_WORKERS", "32").strip()
    try:
        value = int(raw)
    except Exception:
        value = 32
    return max(1, min(256, value))


def _clean(s: str) -> str:
    return (s or "").strip()


def _target_intention(epic: EpicModel) -> str:
    # Excel analysis is based on next increment intention first.
    return _clean(epic.intention_next) or _clean(epic.intention_pi)


def _norm_tokens(text: str) -> set[str]:
    return set(
        w
        for w in re.split(r"[^a-zA-Z0-9àâçéèêëîïôûùüÿñæœ]+", (text or "").lower())
        if len(w) >= 4
    )


def _clip_words(text: str, max_words: int) -> str:
    words = re.sub(r"\s+", " ", _clean(text)).split(" ")
    words = [w for w in words if w]
    if len(words) <= max_words:
        return " ".join(words)
    return " ".join(words[:max_words]).rstrip(",;:") + "..."


def _format_impact_points(lines: List[str]) -> str:
    cleaned = [_clean(x) for x in lines if _clean(x)]
    if not cleaned:
        return ""
    return "\n".join([f"- {x}" for x in cleaned])


def _extract_text(resp) -> str:
    try:
        choices = getattr(resp, "choices", None) or []
        if not choices:
            return ""
        first = choices[0]
        text = getattr(first, "text", None)
        if isinstance(text, str) and text.strip():
            return text.strip()
        message = getattr(first, "message", None)
        if not message:
            return ""
        content = getattr(message, "content", None)
        if isinstance(content, str):
            return content.strip()
        if isinstance(content, list):
            parts: List[str] = []
            for item in content:
                if isinstance(item, str):
                    parts.append(item)
                elif isinstance(item, dict):
                    txt = item.get("text")
                    if isinstance(txt, str):
                        parts.append(txt)
                    elif item.get("type") == "output_text" and isinstance(item.get("value"), str):
                        parts.append(item.get("value"))
                else:
                    txt_attr = getattr(item, "text", None)
                    if isinstance(txt_attr, str):
                        parts.append(txt_attr)
            return "".join(parts).strip()
    except Exception:
        return ""
    return ""


def _parse_json_payload(raw: str) -> dict | None:
    text = _clean(raw)
    if not text:
        return None
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"\{.*\}", text, flags=re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except Exception:
        return None


def _fallback_analysis(entry: EpicEntry) -> EpicAnalysis:
    epic = entry.epic
    desc = _clean(epic.description)
    intention = _target_intention(epic)
    features = [_clean(f) for f in epic.features if _clean(f)]
    base_text = desc or intention or epic.name
    impact_lines: List[str] = []
    if intention:
        impact_lines.append(f"Impact cible: {_clip_words(intention, 24)}")
    if features:
        impact_lines.append(f"Levier delivery: {len(features)} feature(s) identifiee(s) pour cet epic.")
    else:
        impact_lines.append("Levier delivery: aucune feature PI renseignee a ce stade.")
    impact_lines.append("Effet attendu: meilleure alignement entre intention produit et execution PI.")

    synthese_head = _clip_words(base_text, 60) if base_text else "Synthese indisponible."
    synthese = f"{synthese_head}\nImpacts cles:\n{_format_impact_points(impact_lines)}"
    intention_short = _clip_words(intention or "Intention non renseignee.", INTENTION_MAX_WORDS)

    coverage = "insuffisante"
    justification = "Aucune verification possible."
    if features and intention:
        inter = _norm_tokens(" ".join(features)).intersection(_norm_tokens(intention))
        if len(inter) >= 4:
            coverage = "couvre"
            justification = "Les features couvrent globalement les mots-cles de l'intention."
        elif len(inter) >= 1:
            coverage = "partielle"
            justification = "Couverture partielle; certains aspects de l'intention manquent."
        else:
            coverage = "insuffisante"
            justification = "Peu de recouvrement entre intention et features."
    elif features and not intention:
        coverage = "partielle"
        justification = "Des features existent mais l'intention est absente ou trop vague."
    elif not features and intention:
        coverage = "insuffisante"
        justification = "Aucune feature PI pour couvrir l'intention."

    proposals = [
        (f"{epic.name}: formaliser un parcours cible utilisateur", "Clarification des besoins et reduction des retours."),
        (f"{epic.name}: definir des criteres de succes mesurables", "Pilotage des resultats et meilleure priorisation."),
    ]
    return EpicAnalysis(
        synthese_courte=_clip_words(synthese, SYNTH_MAX_WORDS),
        intention_courte=intention_short,
        couverture_features=coverage,
        justification=_clip_words(justification, JUSTIF_MAX_WORDS),
        proposals=proposals,
        ai_used=False,
    )


def _llm_analysis(entry: EpicEntry) -> EpicAnalysis | None:
    if not _llm_enabled():
        return None
    try:
        from openai import OpenAI
    except Exception:
        return None

    epic = entry.epic
    api_key = os.getenv("SCW_SECRET_KEY_LLM")
    base_url = os.getenv("SCW_LLM_BASE_URL", DEFAULT_SCW_BASE_URL)
    model = os.getenv("EXCEL_LLM_MODEL", DEFAULT_EXCEL_MODEL)

    description = _clean(epic.description)
    intention = _target_intention(epic)
    features = [_clean(f) for f in epic.features if _clean(f)]
    features_block = "\n".join([f"- {f}" for f in features[:25]]) if features else "- aucune feature PI"

    prompt = (
        "Tu es un analyste produit. Reponds strictement en JSON.\n"
        "Objectif:\n"
        "1) Produire une synthese claire de l'epic, plus developpee (6 a 10 lignes)\n"
        "2) Dire si les features couvrent l'intention (couvre|partielle|insuffisante)\n"
        "3) Proposer 2 a 4 nouvelles features reformulees avec gains esperes\n"
        "4) Donner 3 a 5 points d'impact concrets (valeur usager/metier, qualite, delai, risque)\n\n"
        "Format JSON strict:\n"
        "{\n"
        '  "synthese_courte": "...",\n'
        '  "intention_courte": "...",\n'
        '  "couverture_features": "couvre|partielle|insuffisante",\n'
        '  "justification": "...",\n'
        '  "impacts_cles": ["...", "..."],\n'
        '  "features_proposees": [\n'
        '    {"feature": "...", "gain": "..."}\n'
        "  ]\n"
        "}\n\n"
        f"Equipe: {entry.team_name}\n"
        f"Epic ID: {epic.id}\n"
        f"Titre: {epic.name}\n"
        f"Description: {description or 'non renseignee'}\n"
        f"Intention: {intention or 'non renseignee'}\n"
        f"Features PI:\n{features_block}\n"
    )

    client = OpenAI(base_url=base_url, api_key=api_key)
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "Tu reponds strictement en JSON valide."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=1600,
            stream=False,
            response_format={"type": "text"},
        )
        raw = _extract_text(resp)
        if not raw:
            # Retry without response_format for provider compatibility.
            resp2 = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "Tu reponds strictement en JSON valide."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.2,
                max_tokens=1600,
                stream=False,
            )
            raw = _extract_text(resp2)
        payload = _parse_json_payload(raw)
        if not payload:
            return None
        synth = _clean(str(payload.get("synthese_courte", "")))
        inten = _clean(str(payload.get("intention_courte", "")))
        coverage = _clean(str(payload.get("couverture_features", ""))).lower()
        justif = _clean(str(payload.get("justification", "")))
        impacts_raw = payload.get("impacts_cles", [])
        impacts: List[str] = []
        if isinstance(impacts_raw, list):
            for it in impacts_raw:
                txt = _clean(str(it))
                if txt:
                    impacts.append(txt)
        props_raw = payload.get("features_proposees", [])
        proposals: List[Tuple[str, str]] = []
        if isinstance(props_raw, list):
            for p in props_raw:
                if not isinstance(p, dict):
                    continue
                feat = _clean(str(p.get("feature", "")))
                gain = _clean(str(p.get("gain", "")))
                if feat or gain:
                    proposals.append((feat, gain))
        if coverage not in {"couvre", "partielle", "insuffisante"}:
            coverage = "partielle"
        if not synth:
            return None
        if not inten:
            inten = "Intention non renseignee."
        if not justif:
            justif = "Justification non fournie."
        if impacts:
            justif = f"{justif}\nImpacts cles:\n{_format_impact_points(impacts[:5])}"
        if not proposals:
            proposals = [
                ("Completer le backlog sur les angles non couverts", "Meilleure couverture de l'intention produit."),
            ]
        return EpicAnalysis(
            synthese_courte=f"{_clip_words(synth, SYNTH_MAX_WORDS)} {ROBOT_TAG}",
            intention_courte=f"{_clip_words(inten, INTENTION_MAX_WORDS)} {ROBOT_TAG}",
            couverture_features=coverage,
            justification=f"{_clip_words(justif, JUSTIF_MAX_WORDS)} {ROBOT_TAG}",
            proposals=[
                (
                    f"{_clip_words(f, FEATURE_PROPOSAL_MAX_WORDS)} {ROBOT_TAG}",
                    f"{_clip_words(g, FEATURE_PROPOSAL_MAX_WORDS)} {ROBOT_TAG}",
                )
                for f, g in proposals[:4]
            ],
            ai_used=True,
        )
    except Exception:
        return None


def _iter_epics(model: BuiltModel) -> List[EpicEntry]:
    entries: List[EpicEntry] = []
    seen: set[int] = set()
    for team in model.teams:
        for epic in team.epics:
            if epic.id in seen:
                continue
            seen.add(epic.id)
            entries.append(EpicEntry(team_name=team.name, epic=epic))
    for epic in model.separate_epics:
        if epic.id in seen:
            continue
        seen.add(epic.id)
        entries.append(EpicEntry(team_name="Epic separee", epic=epic))
    return entries


def generate_epics_excel(model: BuiltModel, out_path: str) -> None:
    entries = _iter_epics(model)
    if not entries:
        wb = Workbook()
        ws = wb.active
        ws.title = "Synthese_Epics"
        ws.append(["Aucune donnee epic disponible"])
        wb.save(out_path)
        return

    analyses: Dict[int, EpicAnalysis] = {}
    lock = threading.Lock()
    stats = {"done": 0, "total": len(entries), "llm_ok": 0, "fallback": 0, "tick": 0}
    is_tty = os.getenv("LLM_LOG_MODE", "compact").strip().lower() != "quiet" and os.isatty(1)

    def _analyze(entry: EpicEntry) -> Tuple[int, EpicAnalysis]:
        llm = _llm_analysis(entry)
        analysis = llm if llm else _fallback_analysis(entry)
        with lock:
            stats["done"] += 1
            if llm is not None:
                stats["llm_ok"] += 1
            else:
                stats["fallback"] += 1
            frames = ["|", "/", "-", "\\"]
            frame = frames[stats["tick"] % len(frames)]
            stats["tick"] += 1
            msg = (
                f"[LLM][Excel] {frame} {stats['done']}/{stats['total']} "
                f"| llm_ok={stats['llm_ok']} fallback={stats['fallback']} | epic={entry.epic.id}"
            )
            if is_tty:
                print(f"\r\033[K{msg}", end="", flush=True)
            else:
                print(msg)
        return entry.epic.id, analysis

    max_workers = min(_excel_llm_max_workers(), len(entries))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(_analyze, e) for e in entries]
        for future in as_completed(futures):
            eid, analysis = future.result()
            analyses[eid] = analysis
    if is_tty:
        print()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Synthese_Epics"
    ws2 = wb.create_sheet("Propositions_Features")

    headers1 = [
        "Equipe",
        "Epic_ID",
        "Epic_Titre",
        "Description_Originale",
        "Intention_Prochain_Increment_Originale",
        "Synthese_Courte",
        "Intention_Courte",
        "Couverture_Features",
        "Justification_Couverture",
        "Nb_Features",
        "Features_PI",
    ]
    ws1.append(headers1)
    headers2 = [
        "Equipe",
        "Epic_ID",
        "Epic_Titre",
        "Feature_Proposee",
        "Gain_Espere",
        "Source",
    ]
    ws2.append(headers2)

    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        epic = entry.epic
        analysis = analyses.get(epic.id) or _fallback_analysis(entry)
        original_intention = _target_intention(epic)
        features_text = "\n".join([f"- {f}" for f in epic.features]) if epic.features else ""
        ws1.append(
            [
                entry.team_name,
                epic.id,
                epic.name,
                _clean(epic.description),
                original_intention,
                analysis.synthese_courte,
                analysis.intention_courte,
                analysis.couverture_features,
                analysis.justification,
                len(epic.features),
                features_text,
            ]
        )
        source = f"LLM {ROBOT_TAG}" if analysis.ai_used else "Fallback local"
        for feat, gain in analysis.proposals:
            ws2.append([entry.team_name, epic.id, epic.name, feat, gain, source])

    for ws in (ws1, ws2):
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths1 = {
        "A": 26, "B": 10, "C": 38, "D": 50, "E": 42, "F": 42, "G": 36, "H": 20, "I": 42, "J": 12, "K": 46
    }
    for col, w in widths1.items():
        ws1.column_dimensions[col].width = w
    widths2 = {"A": 26, "B": 10, "C": 36, "D": 45, "E": 40, "F": 14}
    for col, w in widths2.items():
        ws2.column_dimensions[col].width = w

    wb.save(out_path)
